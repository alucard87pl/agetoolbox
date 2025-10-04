#!/usr/bin/env python3
"""
Flask App for AGE Toolbox - Serves React Frontend + API
"""

import os
import random
import json
from flask import Flask, send_from_directory, jsonify, request
from flask_cors import CORS
import openpyxl

def load_stunts_data():
    """Load stunts data from Excel file"""
    try:
        excel_file = os.path.join(os.path.dirname(__file__), 'data', 'stunts.xlsx')
        if not os.path.exists(excel_file):
            print(f"Warning: Excel file not found: {excel_file}")
            return []
        
        wb = openpyxl.load_workbook(excel_file)
        stunts = []
        stunt_id = 1
        
        # Process each worksheet (category)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            category = sheet_name
            
            # Skip header row, process data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[1] or not row[2]:  # Skip empty rows
                    continue
                
                sp_cost = row[0]
                name = row[1]
                description = row[2]
                setting = row[3] if len(row) > 3 and row[3] else None
                
                stunt = {
                    'id': stunt_id,
                    'name': name,
                    'cost': sp_cost,
                    'category': category,
                    'setting': setting,
                    'description': description
                }
                stunts.append(stunt)
                stunt_id += 1
        
        wb.close()
        return stunts
        
    except Exception as e:
        print(f"Warning: Could not load stunts data from Excel: {e}")
        return []

def add_stunt_to_excel(category, name, cost, description, setting=None):
    """Add a new stunt to the Excel file"""
    excel_file = os.path.join(os.path.dirname(__file__), 'data', 'stunts.xlsx')
    
    try:
        # Load existing workbook or create new one
        if os.path.exists(excel_file):
            wb = openpyxl.load_workbook(excel_file)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
        
        # Get or create worksheet for the category
        if category in wb.sheetnames:
            ws = wb[category]
        else:
            ws = wb.create_sheet(title=category)
            # Add headers if new sheet
            headers = ['SP Cost', 'Name', 'Description', 'Setting']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
                cell.fill = openpyxl.styles.PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        # Find next empty row
        next_row = ws.max_row + 1
        
        # Add the stunt data
        ws.cell(row=next_row, column=1, value=cost)
        ws.cell(row=next_row, column=2, value=name)
        ws.cell(row=next_row, column=3, value=description)
        ws.cell(row=next_row, column=4, value=setting)  # Store blank for Universal stunts
        
        # Style the row
        for col in range(1, 5):
            cell = ws.cell(row=next_row, column=col)
            cell.alignment = openpyxl.styles.Alignment(vertical='top', wrap_text=True)
            if col == 3:  # Description column
                cell.alignment = openpyxl.styles.Alignment(vertical='top', wrap_text=True, horizontal='left')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Set specific widths
        ws.column_dimensions['A'].width = 8   # SP Cost
        ws.column_dimensions['B'].width = 20  # Name
        ws.column_dimensions['C'].width = 50  # Description
        ws.column_dimensions['D'].width = 12  # Setting
        
        # Save the workbook
        wb.save(excel_file)
        wb.close()
        return True
        
    except Exception as e:
        print(f"Error adding stunt: {e}")
        return False

def update_stunt_in_excel(stunt_id, category, name, cost, description, setting=None):
    """Update an existing stunt in the Excel file"""
    excel_file = os.path.join(os.path.dirname(__file__), 'data', 'stunts.xlsx')
    
    try:
        if not os.path.exists(excel_file):
            return False
        
        wb = openpyxl.load_workbook(excel_file)
        current_id = 1
        
        # Find the stunt by ID
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in range(2, ws.max_row + 1):
                if current_id == stunt_id:
                    # Update the stunt data
                    ws.cell(row=row, column=1, value=cost)
                    ws.cell(row=row, column=2, value=name)
                    ws.cell(row=row, column=3, value=description)
                    ws.cell(row=row, column=4, value=setting)  # Store blank for Universal stunts
                    
                    # Save and return
                    wb.save(excel_file)
                    wb.close()
                    return True
                
                # Only increment if row has data
                if ws.cell(row=row, column=1).value:
                    current_id += 1
        
        wb.close()
        return False
        
    except Exception as e:
        print(f"Error updating stunt: {e}")
        return False

def delete_stunt_from_excel(stunt_id):
    """Delete a stunt from the Excel file"""
    excel_file = os.path.join(os.path.dirname(__file__), 'data', 'stunts.xlsx')
    
    try:
        if not os.path.exists(excel_file):
            return False
        
        wb = openpyxl.load_workbook(excel_file)
        current_id = 1
        
        # Find the stunt by ID
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in range(2, ws.max_row + 1):
                if current_id == stunt_id:
                    # Delete the row
                    ws.delete_rows(row)
                    
                    # Save and return
                    wb.save(excel_file)
                    wb.close()
                    return True
                
                # Only increment if row has data
                if ws.cell(row=row, column=1).value:
                    current_id += 1
        
        wb.close()
        return False
        
    except Exception as e:
        print(f"Error deleting stunt: {e}")
        return False

app = Flask(__name__, static_folder='frontend/dist', static_url_path='')
CORS(app)  # Enable CORS for React frontend

# Serve React App
@app.route('/')
def serve_react_app():
    """Serve the React frontend"""
    return send_from_directory('frontend/dist', 'index.html')

@app.route('/<path:path>')
def serve_react_static(path):
    """Serve React static files"""
    try:
        return send_from_directory('frontend/dist', path)
    except:
        # If file not found, serve React app (for client-side routing)
        return send_from_directory('frontend/dist', 'index.html')

# API Routes
@app.route('/api/roll_dice', methods=['POST'])
def api_roll_dice():
    """API endpoint for dice rolling"""
    data = request.get_json()
    bonus = data.get('bonus', 0)
    target = data.get('target', None)
    
    # Roll 2 blue dice and 1 red die (stunt die)
    blue_dice = [random.randint(1, 6) for _ in range(2)]
    red_die = random.randint(1, 6)
    
    # Calculate total and stunt points
    total = sum(blue_dice) + red_die + bonus
    
    success = None
    if target is not None:
        success = total >= target
    
    stunt_points = 0
    has_doubles = False
    
    # Check for doubles (any two dice, including the red die)
    all_dice = blue_dice + [red_die]
    if len(set(all_dice)) < len(all_dice):  # If there are duplicates
        has_doubles = True
        # Only get stunt points if the roll was successful
        if success:
            stunt_points = red_die  # Stunt points are the value of the red die

    response = {
        'blue_dice': blue_dice,
        'red_die': red_die,
        'total': total,
        'bonus': bonus,
        'stunt_points': stunt_points,
        'has_doubles': has_doubles,
        'target': target,
        'success': success,
        'display': f"[{blue_dice[0]}, {blue_dice[1]}] [{red_die}] + {bonus} = {total}"
    }
    return jsonify(response)

@app.route('/api/test', methods=['GET'])
def api_test():
    """Test API endpoint"""
    return jsonify({"status": "ok", "message": "Flask API is working!"})

@app.route('/api/stunts', methods=['GET'])
def api_get_stunts():
    """API endpoint for getting stunt data"""
    stunts_data = load_stunts_data()
    return jsonify(stunts_data)

@app.route('/api/stunts', methods=['POST'])
def api_add_stunt():
    """API endpoint for adding a new stunt"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['name', 'cost', 'category', 'description']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Add the stunt to Excel file
        success = add_stunt_to_excel(
            category=data['category'],
            name=data['name'],
            cost=data['cost'],
            description=data['description'],
            setting=data.get('setting')
        )
        
        if success:
            return jsonify({'message': 'Stunt added successfully'}), 201
        else:
            return jsonify({'error': 'Failed to add stunt'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stunts/<int:stunt_id>', methods=['PUT'])
def api_update_stunt(stunt_id):
    """API endpoint for updating an existing stunt"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['name', 'cost', 'category', 'description']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Update the stunt in Excel file
        success = update_stunt_in_excel(
            stunt_id=stunt_id,
            category=data['category'],
            name=data['name'],
            cost=data['cost'],
            description=data['description'],
            setting=data.get('setting')
        )
        
        if success:
            return jsonify({'message': 'Stunt updated successfully'}), 200
        else:
            return jsonify({'error': 'Stunt not found or update failed'}), 404
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stunts/<int:stunt_id>', methods=['DELETE'])
def api_delete_stunt(stunt_id):
    """API endpoint for deleting a stunt"""
    try:
        success = delete_stunt_from_excel(stunt_id)
        
        if success:
            return jsonify({'message': 'Stunt deleted successfully'}), 200
        else:
            return jsonify({'error': 'Stunt not found or delete failed'}), 404
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    print("Starting Flask app with React frontend...")
    print("React will be served from /")
    print("API endpoints available at /api/*")
    app.run(host='127.0.0.1', port=5000, debug=False)
